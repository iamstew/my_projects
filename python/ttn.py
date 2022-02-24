from openpyxl import load_workbook

try:
	wb = load_workbook('РЕЕСТР.xlsx')
	sheet1 = wb['Лист1']
	sheet2 = wb['Лист2']
except:	
	print("В директории отсутствует файл РЕЕСТР.xlsx\nПерезапустите программу")
	input()

try:
	wb_ttn = load_workbook('ТТН_0.xlsx')
	sheet_ttn1 = wb_ttn['Лист1']
except:
	print("В директории отсутствует файл TTH_0.xlsx\nПерезапустите программу")
	input()

try:
	for row in sheet1.iter_rows(min_row=12, values_only=True):
		k = 0
		if row[0]!=None:
			for col in sheet2.iter_cols(max_row=1, min_col=2, values_only=True):
				sheet_ttn1[col[0]] = row[k+1]
				k +=1
			wb_ttn.save('ttn_'+str(row[0])+'.xlsx')
			print('ttn'+str(row[0]))
except:
	print("Кол-во колонок реестра и описания ячеек не совпадает\nИсправьте, и повторите попытку")
finally:
	print("Программа завершена")
